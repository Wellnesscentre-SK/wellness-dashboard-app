"""Excel / CSV / PDF exports built from a period's merged data."""

from __future__ import annotations

import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from wellness.models import Period
from wellness.services.reports.ppt import (
    CONCERNS, CONCERN_FIELDS, MODE_FIELDS, MODES, REFERRAL_FIELDS,
    REFERRAL_LABELS_SHORT as REFERRALS,
    STAKEHOLDERS, STAKE_FIELDS, VERTICALS, VERTICAL_LABELS, index_rows,
    CONCERN_LABELS_SHORT, STAKE_LABELS_SHORT,
)


def build_excel(period: Period, previous: Period | None = None) -> bytes:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # The worksheet follows the supplied manual-report grid: A:B are the
    # section/team labels, C is Total Cases, and D:AE are the five groups.
    teams = ["WLN Ctr", "Team A", "Your Dost", "Myndwell"]
    groups = [
        ("Gender", ["Male", "Female", "Others / Not to say"],
         ["gender_male", "gender_female", "gender_other"], "D9E2F3"),
        ("Mode of Session", ["Online", "In person", "Phone"],
         ["mode_online", "mode_in_person", "mode_phone"], "E7E6E6"),
        ("Referral type", ["Self", "Director / Kushal Calls", "Dean / HoD / Faculty",
                           "Friend / Family", "Mitr / Saathi"],
         ["referral_self", "referral_director", "referral_dean", "referral_friend",
          "referral_mitr"], "E4DFEC"),
        ("Range of concern addressed", [str(i) for i in range(1, 10)],
         CONCERN_FIELDS, "FCE4D6"),
        ("Stakeholder", [str(i) for i in range(1, 9)], STAKE_FIELDS, "F2F2F2"),
    ]
    all_fields = [field for _name, _labels, fields, _color in groups for field in fields]
    last_column = 3 + len(all_fields)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    vertical = Alignment(horizontal="center", vertical="center", wrap_text=True, textRotation=90)
    border = Border(
        left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"),
    )
    black_bold = Font(name="Calibri", size=9, bold=True, color="000000")
    white_bold = Font(name="Calibri", size=9, bold=True, color="FFFFFF")

    def fill(color):
        return PatternFill("solid", fgColor=color)

    def style_cell(cell, *, fill_color=None, font=None, alignment=center):
        cell.border = border
        cell.alignment = alignment
        if fill_color:
            cell.fill = fill(fill_color)
        if font:
            cell.font = font

    def style_range(row_start, row_end, col_start, col_end, *, fill_color=None,
                   font=None, alignment=center):
        for row in ws.iter_rows(min_row=row_start, max_row=row_end,
                                min_col=col_start, max_col=col_end):
            for cell in row:
                style_cell(cell, fill_color=fill_color, font=font, alignment=alignment)

    # Four header rows: the reference uses a second numeric header row for
    # concern/stakeholder and vertically oriented labels below it.
    ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=2)
    ws["A1"] = "VERTICALS\nNEW / FOLLOW UP"
    style_range(1, 4, 1, 2, fill_color="FFFFFF", font=black_bold)
    ws["A1"].alignment = center

    ws.merge_cells(start_row=1, start_column=3, end_row=4, end_column=3)
    ws["C1"] = "Total Cases"
    style_range(1, 4, 3, 3, fill_color="FCE4D6", font=black_bold, alignment=vertical)

    col = 4
    for group_name, labels, fields, color in groups:
        start = col
        end = col + len(labels) - 1
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        ws.cell(row=1, column=start, value=group_name)
        style_range(1, 1, start, end, fill_color=color, font=black_bold)

        numbered = group_name in ("Range of concern addressed", "Stakeholder")
        for offset, label in enumerate(labels):
            current = start + offset
            if numbered:
                ws.cell(row=2, column=current, value=label)
                style_range(2, 2, current, current, fill_color=color, font=black_bold)
                ws.merge_cells(start_row=3, start_column=current, end_row=4, end_column=current)
                ws.cell(row=3, column=current, value=(
                    CONCERN_LABELS_SHORT[offset] if group_name == "Range of concern addressed"
                    else STAKE_LABELS_SHORT[offset]
                ))
                style_range(3, 4, current, current, fill_color=color, font=black_bold,
                            alignment=vertical)
            else:
                ws.merge_cells(start_row=2, start_column=current, end_row=4, end_column=current)
                ws.cell(row=2, column=current, value=label)
                style_range(2, 4, current, current, fill_color=color, font=black_bold,
                            alignment=vertical)
        col = end + 1

    # Use the latest saved raw sub-team row for each key. Raw rows preserve
    # WLN Ctr and Team A separately even though CaseRow.WC is their aggregate.
    raw_by_key = {}
    raw_rows = list(period.raw_rows.all().order_by("created_at", "id"))
    for raw in raw_rows:
        raw_by_key[f"{raw.case_type}_{raw.sub_team}"] = raw.raw_payload or {}

    merged_rows = index_rows(period)

    def payload_for(case_type, team):
        payload = raw_by_key.get(f"{case_type}_{team}")
        if payload is not None:
            return payload
        # A period created through an older/manual path may have only merged
        # CaseRows. Keep the worksheet useful and never hide WC in Team A.
        if not any(key.startswith(f"{case_type}_") for key in raw_by_key):
            vertical_key = {"WLN Ctr": "WC", "Team A": "TA",
                            "Your Dost": "YD", "Myndwell": "MW"}[team]
            row = merged_rows.get((case_type, vertical_key)) if vertical_key else None
            if row is not None:
                return {field: getattr(row, field, 0) for field in all_fields}
        return {}

    def row_values(case_type, team):
        payload = payload_for(case_type, team)
        values = [int(payload.get(field, 0) or 0) for field in all_fields]
        # Preserve the source workbook's total when it is present. Manual
        # entries already store a validated derived total in this same field.
        total = int(payload.get("total_cases", 0) or 0)
        return total, values

    team_row_fills = {"WLN Ctr": "F8FBFF", "Team A": "F2F2F2",
                      "Your Dost": "FCE4D6", "Myndwell": "EDE7F6"}

    def write_case_block(case_type, start_row, label, value_font_color):
        block_end = start_row + len(teams) - 1
        ws.merge_cells(start_row=start_row, start_column=1, end_row=block_end, end_column=1)
        ws.cell(row=start_row, column=1, value=label)
        style_range(start_row, block_end, 1, 1, fill_color="FFFFFF",
                    font=Font(name="Calibri", size=9, bold=True, color="C00000"))
        ws.cell(row=start_row, column=1).alignment = center

        totals = [0] * len(all_fields)
        total_cases = 0
        for offset, team in enumerate(teams):
            row = start_row + offset
            total, values = row_values(case_type, team)
            total_cases += total
            totals = [left + right for left, right in zip(totals, values)]
            style_range(row, row, 2, last_column, fill_color=team_row_fills[team],
                        font=Font(name="Calibri", size=9, bold=True, color=value_font_color))
            ws.cell(row=row, column=2, value=team)
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row=row, column=3, value=total)
            ws.cell(row=row, column=3).alignment = center
            for offset2, value in enumerate(values, start=4):
                ws.cell(row=row, column=offset2, value=value)
                ws.cell(row=row, column=offset2).alignment = center

        total_row = block_end + 1
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
        total_label = "NEW" if case_type == "new" else "FOLLOW-UP"
        ws.cell(row=total_row, column=1, value=f"Total no. of cases\n{total_label}")
        total_fill = "31859C" if case_type == "new" else "00A8D6"
        style_range(total_row, total_row, 1, 2, fill_color=total_fill, font=white_bold)
        ws.cell(row=total_row, column=1).alignment = center
        style_range(total_row, total_row, 3, last_column, fill_color="F4B183", font=black_bold)
        ws.cell(row=total_row, column=3, value=total_cases)
        ws.cell(row=total_row, column=3).alignment = center
        for offset, value in enumerate(totals, start=4):
            ws.cell(row=total_row, column=offset, value=value)
            ws.cell(row=total_row, column=offset).alignment = center
        return total_row, total_cases, totals

    new_total_row, new_total, new_values = write_case_block(
        "new", 5, "No. of\nNew\nCases", "0000FF")
    fu_total_row, fu_total, fu_values = write_case_block(
        "followup", new_total_row + 1, "No. of\nFollow-\nup cases", "C00000")
    grand_row = fu_total_row + 1
    ws.merge_cells(start_row=grand_row, start_column=1, end_row=grand_row, end_column=2)
    ws.cell(row=grand_row, column=1, value="Grand Total")
    style_range(grand_row, grand_row, 1, last_column, fill_color="8FC7D4", font=black_bold)
    ws.cell(row=grand_row, column=1).alignment = center
    ws.cell(row=grand_row, column=3, value=new_total + fu_total)
    ws.cell(row=grand_row, column=3).alignment = center
    for offset, value in enumerate([a + b for a, b in zip(new_values, fu_values)], start=4):
        ws.cell(row=grand_row, column=offset, value=value)
        ws.cell(row=grand_row, column=offset).alignment = center

    # The secondary/operations block is intentionally on Summary, immediately
    # below the cross-tab. WC and Team A remain separate throughout.
    secondary_header_row = grand_row + 2
    secondary_label_row = secondary_header_row + 1
    secondary_value_row = secondary_header_row + 2
    secondary_groups = [
        ("Unrecognised", 2, "E4DFEC"),
        ("Total no. of sessions", 5, "FFF2CC"),
        ("Early Prevention warning", 5, "FCE4D6"),
        ("No. of cases which did not turn-up even\nwhen it was advised", 5, "F2F2F2"),
        ("As of date no. of active cases", 5, "E2EFDA"),
        ("Clients undergoing more than 4 sessions in a month", 5, "FCE4D6"),
        ("Mode of enquiries", 3, "FFF2CC"),
    ]
    cursor = 1
    for title, span, color in secondary_groups:
        ws.merge_cells(start_row=secondary_header_row, start_column=cursor,
                       end_row=secondary_header_row, end_column=cursor + span - 1)
        ws.cell(row=secondary_header_row, column=cursor, value=title)
        style_range(secondary_header_row, secondary_header_row, cursor, cursor + span - 1,
                    fill_color=color, font=Font(name="Calibri", size=9, bold=True, color="C00000"))
        cursor += span

    sec_metrics = {metric.vertical: metric for metric in period.secondary_metrics.all()}
    total_metric = sec_metrics.get("Total")
    metric_fields = [
        ("total_sessions", "0000FF"), ("early_prevention_warning", "C00000"),
        ("no_show_turn_up", "C00000"), ("active_cases", "008000"),
        ("clients_over_4_sessions", "0000FF"),
    ]
    # Unrecognised is retained as the two labelled cells from the reference;
    # the current model has no unrecognised metric, so their values are zero.
    for col, team in ((1, "Your Dost"), (2, "Myndwell")):
        ws.cell(row=secondary_label_row, column=col, value=team)
        style_cell(ws.cell(row=secondary_label_row, column=col), fill_color="E4DFEC",
                   font=black_bold, alignment=vertical)
        ws.cell(row=secondary_value_row, column=col, value=0)
        style_cell(ws.cell(row=secondary_value_row, column=col), fill_color="FFFFFF",
                   font=black_bold)

    cursor = 3
    for field, font_color in metric_fields:
        values = []
        for team in ("WC", "TA", "YD", "MW", "Total"):
            if team == "Total":
                value = getattr(total_metric, field, 0) if total_metric else 0
            else:
                metric = sec_metrics.get(team)
                value = getattr(metric, field, 0) if metric else 0
            values.append(value)
        for offset, (team, value) in enumerate(zip(("WLN Ctr", "Team A", "Your Dost", "Myndwell", "Total"), values)):
            column = cursor + offset
            ws.cell(row=secondary_label_row, column=column, value=team)
            style_cell(ws.cell(row=secondary_label_row, column=column), fill_color="E6E6FA",
                       font=Font(name="Calibri", size=9, bold=True, color=font_color),
                       alignment=vertical)
            ws.cell(row=secondary_value_row, column=column, value=value)
            style_cell(ws.cell(row=secondary_value_row, column=column), fill_color="FFFFFF",
                       font=black_bold)
        cursor += 5

    enquiry = getattr(period, "enquiry_modes", None)
    for offset, (label, value) in enumerate((
        ("Thro Mail", getattr(enquiry, "mail", 0) if enquiry else 0),
        ("Thro Calls", getattr(enquiry, "calls_recd", 0) if enquiry else 0),
        ("Thro Calls\nOutgoing", getattr(enquiry, "calls_out", 0) if enquiry else 0),
    )):
        column = cursor + offset
        ws.cell(row=secondary_label_row, column=column, value=label)
        style_cell(ws.cell(row=secondary_label_row, column=column), fill_color="E6E6FA",
                   font=black_bold, alignment=vertical)
        ws.cell(row=secondary_value_row, column=column, value=value)
        style_cell(ws.cell(row=secondary_value_row, column=column), fill_color="FFFFFF",
                   font=black_bold)

    # Apply borders to every visible cell, including the merged-label ranges.
    for row in range(1, secondary_value_row + 1):
        for column in range(1, last_column + 1):
            ws.cell(row=row, column=column).border = border

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 46
    ws.row_dimensions[4].height = 46
    for row in range(5, grand_row + 1):
        ws.row_dimensions[row].height = 24
    ws.row_dimensions[secondary_header_row].height = 32
    ws.row_dimensions[secondary_label_row].height = 84
    ws.row_dimensions[secondary_value_row].height = 24
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 8
    for column in range(4, last_column + 1):
        ws.column_dimensions[get_column_letter(column)].width = 6
    ws.freeze_panes = "D5"
    ws.sheet_view.zoomScale = 70
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Keep the audit/verification sheets below compatible with the historical
    # four-team names used by the importer.
    TEAMS = teams
    ws3 = wb.create_sheet("Raw Rows")
    def _write_table(ws_sheet, start_row, title, headers, data):
        ws_sheet.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
        r = start_row + 1
        for c_idx, h in enumerate(headers, 1):
            cell = ws_sheet.cell(row=r, column=c_idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F46E5")
        for row_data in data:
            r += 1
            for c_idx, v in enumerate(row_data, 1):
                ws_sheet.cell(row=r, column=c_idx, value=v)
        return r + 1

    _write_table(ws3, 1, "Raw Sub-team Rows", ["Entry", "Case Type", "Sub-team", "Source",
                                               "Needs Review", "Reason"],
                 [[r.entry_no, r.case_type, r.sub_team, r.source, r.needs_review,
                   r.reason or ""] for r in period.raw_rows.all()])

    # Verification Reference Sheet (Spec 15-22)
    ws_v = wb.create_sheet("Verification Reference")
    ws_v.cell(row=1, column=1, value="Verification reference").font = Font(bold=True, size=14)

    v_headers = ["Row", "Gender", "Session", "Referral", "Concern", "Stakeholder"]
    for c_idx, h in enumerate(v_headers, 1):
        cell = ws_v.cell(row=2, column=c_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E293B")

    v_rows = []
    for ct, ct_label in [("new", "New"), ("followup", "Follow-up")]:
        for t in TEAMS:
            v_rows.append((f"{ct_label} — {t}", ct, t))
        v_rows.append((f"Total {ct_label}", ct, "TOTAL"))
    v_rows.append(("Grand Total", "ALL", "GRAND_TOTAL"))

    def _calc_row_payload(ct, team_key):
        if team_key not in ("TOTAL", "GRAND_TOTAL"):
            return raw_by_key.get(f"{ct}_{team_key}", {})
        if team_key == "TOTAL":
            res = {}
            for t in TEAMS:
                p = raw_by_key.get(f"{ct}_{t}", {})
                for k, v in p.items():
                    res[k] = res.get(k, 0) + (v or 0)
            return res
        # GRAND_TOTAL
        res = {}
        for ctype in ("new", "followup"):
            for t in TEAMS:
                p = raw_by_key.get(f"{ctype}_{t}", {})
                for k, v in p.items():
                    res[k] = res.get(k, 0) + (v or 0)
        return res

    fill_true = PatternFill("solid", fgColor="D1FAE5")
    fill_false = PatternFill("solid", fgColor="FEE2E2")
    font_true = Font(color="065F46", bold=True)
    font_false = Font(color="991B1B", bold=True)

    for idx, (label, ct, team_key) in enumerate(v_rows, start=3):
        p = _calc_row_payload(ct, team_key)
        tot = (p.get("gender_male", 0) + p.get("gender_female", 0) + p.get("gender_other", 0))

        g_valid = True # Total Cases derived from Gender
        s_valid = tot == (p.get("mode_online", 0) + p.get("mode_in_person", 0) + p.get("mode_phone", 0))
        r_valid = tot == (p.get("referral_self", 0) + p.get("referral_director", 0) + p.get("referral_dean", 0) + p.get("referral_friend", 0) + p.get("referral_mitr", 0))
        c_valid = tot == sum(p.get(f, 0) for f in [
            "concern_anxiety", "concern_stress", "concern_career", "concern_interpersonal",
            "concern_self_dev", "concern_clinical", "concern_addiction", "concern_medical", "concern_suicidal"
        ])
        st_valid = tot == sum(p.get(f, 0) for f in [
            "stake_ug", "stake_pg", "stake_phd", "stake_dual", "stake_faculty",
            "stake_employee_family", "stake_postdoc", "stake_unidentified"
        ])

        ws_v.cell(row=idx, column=1, value=label).font = Font(bold=True)
        for col_i, val in enumerate([g_valid, s_valid, r_valid, c_valid, st_valid], start=2):
            cell = ws_v.cell(row=idx, column=col_i, value="TRUE" if val else "FALSE")
            cell.fill = fill_true if val else fill_false
            cell.font = font_true if val else font_false

    for i in range(1, 7):
        ws_v.column_dimensions[get_column_letter(i)].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()



def _per_vertical(rows, field):
    """Sum field across new+followup for each of the 4 verticals, plus grand total."""
    vals = []
    for v in VERTICALS:
        val = getattr(rows[("new", v)], field, 0) + getattr(rows[("followup", v)], field, 0)
        vals.append(val)
    vals.append(sum(vals))
    return vals


def build_csv(period: Period, previous: Period | None = None) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    rows = index_rows(period)

    w.writerow([f"{period.report_type} report,{period.period_start},{period.period_end}"])
    w.writerow([])
    w.writerow(["case_type", "vertical", "total_cases", "gender_male", "gender_female", "gender_other",
                "mode_online", "mode_in_person", "mode_phone", "needs_review"])
    for ct in ("new", "followup"):
        for v in VERTICALS:
            r = rows[(ct, v)]
            w.writerow([ct, v, r.total_cases, r.gender_male, r.gender_female, r.gender_other,
                        r.mode_online, r.mode_in_person, r.mode_phone, r.needs_review])
    if previous:
        w.writerow([])
        w.writerow([f"PREVIOUS {previous.report_type} report,{previous.period_start},{previous.period_end}"])
        for ct in ("new", "followup"):
            for v in VERTICALS:
                r = index_rows(previous)[(ct, v)]
                w.writerow([ct, v, r.total_cases, r.gender_male, r.gender_female, r.gender_other,
                            r.mode_online, r.mode_in_person, r.mode_phone, r.needs_review])
    return buf.getvalue().encode("utf-8-sig")


def build_pdf(period: Period, previous: Period | None = None) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    rows = index_rows(period)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"Weekly Wellness Report" if period.report_type == "weekly"
                  else "Monthly Wellness Report", styles["Title"]),
        Spacer(1, 6),
        Paragraph(f"{period.period_start} to {period.period_end}", styles["Normal"]),
        Spacer(1, 12),
    ]

    headers = ["Case Type", "Vertical", "Total", "M", "F", "O"]
    data = [headers]
    for ct in ("new", "followup"):
        for v in ("WC", "TA", "YD", "MW"):
            r = rows[(ct, v)]
            data.append([ct.capitalize(), v, r.total_cases, r.gender_male,
                         r.gender_female, r.gender_other])
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (2, 1), (-1, -1), "CENTER"),
    ]))
    story.append(t)
    story.append(Spacer(1, 18))

    def _cat_table(title, names, fields):
        d = [["Category", "WC", "Team A", "YD", "MW", "Total"]]
        for name, field in zip(names, fields):
            d.append([name] + _per_vertical(rows, field))
        tt = Table(d, repeatRows=1)
        tt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0EA5E9")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(Paragraph(title, styles["Heading2"]))
        story.append(tt)
        story.append(Spacer(1, 12))

    _cat_table("Concerns", CONCERNS, CONCERN_FIELDS)
    _cat_table("Stakeholders", STAKEHOLDERS, STAKE_FIELDS)
    _cat_table("Referrals", REFERRALS, REFERRAL_FIELDS)
    _cat_table("Session Modes", MODES, MODE_FIELDS)

    doc.build(story)
    return buf.getvalue()


def build_comparison_excel(period_a: Period, period_b: Period, insights: dict) -> bytes:
    """Excel workbook for an AI period-over-period comparison.

    Takes an insights dict from ``wellness.services.insights.compare_periods``.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "AI Comparison"

    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")
    group_fill = PatternFill("solid", fgColor="E0E7FF")
    pos_fill = PatternFill("solid", fgColor="D1FAE5")
    neg_fill = PatternFill("solid", fgColor="FEE2E2")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    a = insights["period_a"]
    b = insights["period_b"]
    t = insights["totals"]

    ws.cell(row=1, column=1, value="AI DATA ANALYSIS").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"{insights['comparison_label']}").font = bold
    ws.cell(row=3, column=1, value=f"{a['label']}  vs  {b['label']}").font = Font(italic=True, color="64748B")

    # Summary table
    r = 5
    headers = ["Metric", a["label"], b["label"], "Change", "% Change"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = white_bold
        cell.fill = header_fill
        cell.border = border
    r += 1

    def _pct(p):
        return f"{p:+.1f}%" if p is not None else "—"

    summary = [
        ("Total cases", a["total"], b["total"], t["delta_total"], _pct(t["pct_total"])),
        ("New cases", a["new"], b["new"], t["delta_new"], _pct(t["pct_new"])),
        ("Follow-up cases", a["followup"], b["followup"], t["delta_followup"], _pct(t["pct_followup"])),
        ("Total sessions", a["total_sessions"], b["total_sessions"], t["delta_sessions"], _pct(t["pct_sessions"])),
    ]
    for label, av, bv, d, p in summary:
        ws.cell(row=r, column=1, value=label).font = bold
        ws.cell(row=r, column=2, value=av)
        ws.cell(row=r, column=3, value=bv)
        dc = ws.cell(row=r, column=4, value=d)
        ws.cell(row=r, column=5, value=p)
        if d:
            dc.fill = pos_fill if d > 0 else neg_fill
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = border
        r += 1

    # Category deltas
    r += 1
    ws.cell(row=r, column=1, value="Category breakdown — period-over-period").font = Font(bold=True, size=12)
    r += 1
    for group, label in [("gender", "Gender"), ("mode", "Mode of Session"),
                         ("referral", "Referral Type"), ("concern", "Range of Concern"),
                         ("stakeholder", "Stakeholder"), ("vertical", "Vertical")]:
        ws.cell(row=r, column=1, value=label).font = white_bold
        ws.cell(row=r, column=1).fill = group_fill
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1
        for c, h in enumerate(["Dimension", a["label"], b["label"], "Change", "% Change"], 1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = bold
            cell.border = border
        r += 1
        for entry in insights["category_deltas"].get(group, []):
            ws.cell(row=r, column=1, value=entry["label"]).border = border
            ws.cell(row=r, column=2, value=entry["a"]).border = border
            ws.cell(row=r, column=3, value=entry["b"]).border = border
            dc = ws.cell(row=r, column=4, value=entry["delta"])
            dc.border = border
            ws.cell(row=r, column=5, value=_pct(entry["pct"])).border = border
            if entry["delta"]:
                dc.fill = pos_fill if entry["delta"] > 0 else neg_fill
            r += 1
        r += 1

    # AI insights bullets
    ws.cell(row=r, column=1, value="AI insights").font = Font(bold=True, size=12)
    r += 1
    for bullet in insights.get("insights", []):
        cell = ws.cell(row=r, column=1, value=f"•  {bullet['text']}")
        cell.alignment = wrap
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1

    ws.column_dimensions["A"].width = 34
    for col in ("B", "C", "D", "E"):
        ws.column_dimensions[col].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_aggregate_excel(title: str, subtitle: str, merged: dict,
                          source_rows: list | None = None) -> bytes:
    """Clean multi-sheet workbook for COMBINED monthly / yearly reports.
    merged: ppt_generator dict format (see report_center.merge_period_dicts)."""
    header_fill = PatternFill("solid", fgColor="2F5597")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="2F5597")
    sub_font = Font(name="Calibri", size=10, italic=True, color="595959")
    body = Font(name="Calibri", size=10)
    bold = Font(name="Calibri", size=10, bold=True)

    wb = Workbook()

    def sheet(name, headers, rows, widths=None):
        ws = wb.create_sheet(name)
        c = ws.cell(row=1, column=1, value=headers[0] and name.replace("_", " ").title())
        c.font = title_font
        for j, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=j, value=h)
            cell.fill = header_fill
            cell.font = header_font
        for i, row in enumerate(rows, 4):
            for j, v in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=v)
                cell.font = bold if isinstance(row[0], str) and j == 1 else body
        for j, w in enumerate(widths or [30] + [14] * (len(headers) - 1), 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        return ws

    grand = int(merged.get("grand", 0) or 0)
    new_n = int(merged.get("new", 0) or 0)
    fu_n = int(merged.get("followup", 0) or 0)

    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value=title).font = title_font
    ws.cell(row=2, column=1, value=subtitle).font = sub_font
    summary_rows = [
        ("Total Cases", grand, "100%"),
        ("New Cases", new_n, f"{new_n / grand * 100:.0f}%" if grand else "0%"),
        ("Follow-up Cases", fu_n, f"{fu_n / grand * 100:.0f}%" if grand else "0%"),
    ]
    for i, (lbl, val, pct) in enumerate(summary_rows, 4):
        ws.cell(row=i, column=1, value=lbl).font = bold
        ws.cell(row=i, column=2, value=val).font = body
        ws.cell(row=i, column=3, value=pct).font = body
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12

    vert = merged.get("vertical") or {}
    v_names = {"WC": "WLN Ctr", "TA": "Team A", "YD": "Your Dost", "MW": "Myndwell"}
    sheet("Verticals", ["Vertical", "New", "Follow-up", "Total", "% Share"],
          [[v_names.get(k, k), d.get("new", 0), d.get("followup", 0),
            d.get("total", 0),
            f"{d.get('total', 0) / grand * 100:.0f}%" if grand else "0%"]
           for k, d in vert.items()])

    def cat_rows(dim, labels):
        return [[lbl, int(merged.get(dim, {}).get(lbl, 0) or 0),
                 f"{(merged.get(dim, {}).get(lbl, 0) or 0) / grand * 100:.0f}%"
                 if grand else "0%"] for lbl in labels]

    from wellness.services.reports.ppt import GENDERS, MODES, REFERRALS, CONCERNS, STAKEHOLDERS
    sheet("Demographics", ["Category", "Count", "% Share"],
          cat_rows("gender", GENDERS) + cat_rows("mode", MODES)
          + cat_rows("referral", REFERRALS))
    sheet("Concern", ["Concern", "Count", "% Share"], cat_rows("concern", CONCERNS))
    sheet("Stakeholder", ["Stakeholder", "Count", "% Share"],
          cat_rows("stakeholder", STAKEHOLDERS))

    if source_rows:
        sheet("Sources", ["Period", "New", "Follow-up", "Total"],
              [[r["label"], r.get("new", 0), r.get("followup", 0), r.get("total", 0)]
               for r in source_rows], widths=[40, 14, 14, 14])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build(kind: str, period: Period, previous: Period | None = None) -> tuple[str, bytes, str]:
    """Return (filename, bytes, content_type) for the requested export kind."""
    slug = f"{period.report_type}_{period.period_start}_{period.period_end}"
    if kind == "xlsx":
        return f"{slug}.xlsx", build_excel(period, previous), \
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if kind == "csv":
        return f"{slug}.csv", build_csv(period, previous), "text/csv"
    if kind == "pdf":
        return f"{slug}.pdf", build_pdf(period, previous), "application/pdf"
    raise ValueError(f"Unknown export kind: {kind}")
