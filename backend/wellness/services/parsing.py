"""Positional parser for the fixed-layout Wellness Centre Excel report.

Both the weekly and monthly reports share an identical fixed layout
(see spec section 15). This module parses positionally — fixed row/column
ranges plus row-label matching — never generic header inference.
"""

from __future__ import annotations

import datetime as dt
import hashlib
import io
import re
from dataclasses import dataclass, field
from typing import Optional

import openpyxl

# ---------------------------------------------------------------------------
# Layout constants (ground truth from the real files)
# ---------------------------------------------------------------------------

REPORT_TYPES = ("weekly", "monthly")
TEAMS = ("WLN Ctr", "Team A", "Your Dost", "Myndwell")
# Keep the four source categories independent. WC and Team A are distinct
# reporting verticals even though older versions of the app merged them.
VERTICALS = ("WC", "TA", "YD", "MW")

# Case rows: New block rows 6-9, Follow-up block rows 11-14 (sheet row numbers).
NEW_ROWS = {team: 6 + i for i, team in enumerate(TEAMS)}
FOLLOWUP_ROWS = {team: 11 + i for i, team in enumerate(TEAMS)}

# Excel column index of the first data column (D) = 4.
DATA_START_COL = 4  # D

# The 29 numeric data columns, in fixed order (spec section 20).
COLUMN_SCHEMA = [
    ("total_cases", "total"),
    ("gender_male", "gender"),
    ("gender_female", "gender"),
    ("gender_other", "gender"),
    ("mode_online", "mode"),
    ("mode_in_person", "mode"),
    ("mode_phone", "mode"),
    ("referral_self", "referral"),
    ("referral_director", "referral"),
    ("referral_dean", "referral"),
    ("referral_friend", "referral"),
    ("referral_mitr", "referral"),
    ("concern_anxiety", "concern"),
    ("concern_stress", "concern"),
    ("concern_career", "concern"),
    ("concern_interpersonal", "concern"),
    ("concern_self_dev", "concern"),
    ("concern_clinical", "concern"),
    ("concern_addiction", "concern"),
    ("concern_medical", "concern"),
    ("concern_suicidal", "concern"),
    ("stake_ug", "stakeholder"),
    ("stake_pg", "stakeholder"),
    ("stake_phd", "stakeholder"),
    ("stake_dual", "stakeholder"),
    ("stake_faculty", "stakeholder"),
    ("stake_employee_family", "stakeholder"),
    ("stake_postdoc", "stakeholder"),
    ("stake_unidentified", "stakeholder"),
]
COLUMN_NAMES = [name for name, _ in COLUMN_SCHEMA]
ALL_FIELDS = COLUMN_NAMES + [
    "total_sessions",
    "early_prevention_warning",
    "no_show_turn_up",
    "active_cases",
    "clients_over_4_sessions",
    "thro_mail",
    "thro_calls_recd",
    "thro_calls_out",
]

# The 5 live category-sum checks (spec section 2.4). Stakeholder ALWAYS uses
# all 8 columns — the source template's SUM(Y:AE) bug on some rows is NOT
# replicated (spec's explicit instruction).
CHECK_GROUPS = {
    "gender": ["gender_male", "gender_female", "gender_other"],
    "session": ["mode_online", "mode_in_person", "mode_phone"],
    "referral": [
        "referral_self",
        "referral_director",
        "referral_dean",
        "referral_friend",
        "referral_mitr",
    ],
    "concern": [
        "concern_anxiety",
        "concern_stress",
        "concern_career",
        "concern_interpersonal",
        "concern_self_dev",
        "concern_clinical",
        "concern_addiction",
        "concern_medical",
        "concern_suicidal",
    ],
    "stakeholder": [
        "stake_ug",
        "stake_pg",
        "stake_phd",
        "stake_dual",
        "stake_faculty",
        "stake_employee_family",
        "stake_postdoc",
        "stake_unidentified",
    ],
}

# Secondary metrics block (rows 18-20): (row18 group header col, start col of
# teams, total col). Team columns run WLN Ctr, Team A, Your Dost, Myndwell.
SECONDARY_GROUPS = {
    "total_sessions": (18, 4, 8),       # D..G teams, H total
    "early_prevention_warning": (9, 9, 13),   # I..L teams, M total
    "no_show_turn_up": (14, 14, 18),    # N..Q teams, R total
    "active_cases": (19, 19, 23),       # S..V teams, W total
    "clients_over_4_sessions": (24, 24, 28),  # X..AA teams, AB total
}
ENQUIRY_COLS = {"mail": 29, "calls_recd": 30, "calls_out": 31}  # AC..AE
STRAY_COLS = (2, 3)  # B, C — legacy cells (B19='Your Dost', C19='Myndwell')

_MONTHS = {
    m.lower(): i
    for i, m in enumerate(
        [
            "Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
        ],
        start=1,
    )
}

# Error codes (spec section 18)
ERR_SHEET_STRUCTURE = "SHEET_STRUCTURE_UNRECOGNIZED"
ERR_DUPLICATE_PERIOD = "DUPLICATE_PERIOD"
ERR_ROW_VALIDATION = "ROW_VALIDATION_FAILED"
ERR_MISSING_MANDATORY = "MISSING_MANDATORY_FIELD"
ERR_NEGATIVE_VALUE = "NEGATIVE_VALUE"
ERR_NON_INTEGER = "NON_INTEGER_VALUE"
ERR_PARTIAL_PERIOD = "PARTIAL_PERIOD"


# ---------------------------------------------------------------------------
# DTOs
# ---------------------------------------------------------------------------

@dataclass
class CheckResult:
    name: str
    passed: bool
    expected: int
    actual: int

    @property
    def off_by(self) -> int:
        return self.actual - self.expected


@dataclass
class CellIssue:
    field: str
    cell: str  # e.g. "E6"
    code: str
    message: str


@dataclass
class SubTeamRow:
    case_type: str  # 'new' | 'followup'
    sub_team: str
    sheet_row: int
    columns: dict  # field -> int
    checks: list = field(default_factory=list)
    status: str = "ready"  # 'ready' | 'warning' | 'rejected'
    issues: list = field(default_factory=list)

    @property
    def needs_review(self) -> bool:
        return self.status == "warning"

    @property
    def reason(self) -> Optional[str]:
        if self.status == "rejected":
            return "; ".join(i.message for i in self.issues)
        if self.status == "warning":
            failed = [c for c in self.checks if not c.passed]
            return "; ".join(
                f"{c.name} off by {c.off_by}" for c in failed
            )
        return None


@dataclass
class SecondaryMetrics:
    total_sessions: dict = field(default_factory=lambda: {v: 0 for v in VERTICALS} | {"Total": 0})
    early_prevention_warning: dict = field(default_factory=lambda: {v: 0 for v in VERTICALS} | {"Total": 0})
    no_show_turn_up: dict = field(default_factory=lambda: {v: 0 for v in VERTICALS} | {"Total": 0})
    active_cases: dict = field(default_factory=lambda: {v: 0 for v in VERTICALS} | {"Total": 0})
    clients_over_4_sessions: dict = field(default_factory=lambda: {v: 0 for v in VERTICALS} | {"Total": 0})
    enquiry_modes: dict = field(default_factory=lambda: {"mail": 0, "calls_recd": 0, "calls_out": 0})
    stray_cells: dict = field(default_factory=dict)


@dataclass
class ParsedReport:
    report_type: str
    period_start: dt.date
    period_end: dt.date
    title: str
    rows: list
    secondary: SecondaryMetrics
    structure_ok: bool = True
    structure_issues: list = field(default_factory=list)
    warnings: list = field(default_factory=list)
    title_range_mismatch: bool = False
    file_sha256: str = ""

    @property
    def period_label(self) -> str:
        return f"{self.period_start.isoformat()} to {self.period_end.isoformat()}"

    def merged_rows(self) -> dict:
        """(case_type, vertical) -> {field: int}, computed from accepted rows."""
        return merge_verticals(self.rows)


# ---------------------------------------------------------------------------
# Title / date parsing
# ---------------------------------------------------------------------------

_TITLE_RE = re.compile(
    r"^(?P<rtype>Weekly|Monthly)\s+Wellness\s+Report\s+From\s+"
    r"(?P<start>.+?)\s+to\s+(?P<end>.+?)\s*$",
    re.IGNORECASE,
)
_DATE_RE = re.compile(
    r"(?P<day>\d{1,2})(?:st|nd|rd|th)?\s+(?P<month>[A-Za-z]+)"
    r"(?:\s*(?P<year>\d{4}))?\s*$"
)


def parse_title_date(value: str):
    """Parse '29 July', '04th August 2026', '1st Jan' etc. into a date.

    Ordinal suffixes are stripped permissively; a missing year is filled by
    the caller with the other endpoint's year. Returns (date, year_or_None).
    """
    if value is None:
        return None, None
    m = _DATE_RE.match(value.strip())
    if not m:
        return None, None
    day = int(m.group("day"))
    month_name = m.group("month").lower()
    month = _MONTHS.get(month_name[:3])
    if month is None:
        return None, None
    year = int(m.group("year")) if m.group("year") else None
    return dt.date(year or 2000, month, day), year


def parse_title(title: str):
    """Return (report_type, start, end, year) or None if the title is not a
    Wellness report title.

    Start date may omit the month when it shares the end date's month, e.g.
    'From 01st to 30th July 2026' (start = day only).
    """
    m = _TITLE_RE.match(title.strip())
    if not m:
        return None
    rtype = "weekly" if m.group("rtype").lower() == "weekly" else "monthly"
    start_raw = m.group("start").strip()
    end_raw = m.group("end").strip()

    end, end_year = parse_title_date(end_raw)
    if not end:
        return None
    year = end_year or dt.date.today().year

    start = _parse_start_date(start_raw, year, end)
    if not start:
        return None
    return rtype, start, end


def _parse_start_date(raw: str, year: int, end: dt.date):
    """Start date is either a full 'day Month [year]' or a bare day+ordinal
    ('01st') that inherits the end date's month."""
    raw = raw.strip()
    d, y = parse_title_date(raw)
    if d is not None:
        return dt.date(y or year, d.month, d.day)
    m = re.match(r"^(\d{1,2})(?:st|nd|rd|th)?\s*$", raw)
    if m:
        return dt.date(year, end.month, int(m.group(1)))
    return None


# ---------------------------------------------------------------------------
# Cell value handling
# ---------------------------------------------------------------------------

def _coerce_int(value, cell_ref: str, field: str) -> tuple[int, Optional[CellIssue]]:
    """Coerce a cell value to a non-negative integer.

    Blank/None, empty strings, booleans, and non-numeric text are treated as 0.
    Non-integer floats are truncated. Negative values produce an issue.
    Rows should never be rejected simply because a cell is blank.
    """
    if value is None or isinstance(value, bool):
        return 0, None
    if isinstance(value, str):
        s = value.strip()
        if s == "" or s == "-":
            return 0, None
        try:
            n = int(float(s))
            if n < 0:
                return 0, CellIssue(field, cell_ref, ERR_NEGATIVE_VALUE,
                                    f"Row/cell {cell_ref}: negative values aren't allowed.")
            return n, None
        except (ValueError, OverflowError):
            return 0, None
    if isinstance(value, float):
        if value < 0:
            return 0, CellIssue(field, cell_ref, ERR_NEGATIVE_VALUE,
                                f"Row/cell {cell_ref}: negative values aren't allowed.")
        return int(value), None
    if isinstance(value, int):
        if value < 0:
            return 0, CellIssue(field, cell_ref, ERR_NEGATIVE_VALUE,
                                f"Row/cell {cell_ref}: negative values aren't allowed.")
        return value, None
    return 0, None


def _sum_formula_range(formula: str) -> Optional[list[str]]:
    """Extract cell refs from a '=SUM(D20:G20)' formula, or None."""
    m = re.search(r"=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)", formula, re.IGNORECASE)
    if not m:
        return None
    c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
    from openpyxl.utils import column_index_from_string

    col1 = column_index_from_string(c1)
    col2 = column_index_from_string(c2)
    return [f"{openpyxl.utils.get_column_letter(c)}{r}" for r in range(r1, r2 + 1) for c in range(col1, col2 + 1)]


def _detect_team_rows(ws):
    """Scan column C for team labels and return (new_rows, followup_rows, fu_present).

    Each is a dict {team_name: row_number}. If only one block of 4 teams is
    found, followup_rows will be empty and fu_present will be False.
    Rows within a block that have no valid team label are returned with
    sub_team='<missing>'.
    """
    team_occurrences = []
    for row_idx in range(1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=3).value
        if val in TEAMS:
            team_occurrences.append((row_idx, val))

    if not team_occurrences:
        return {}, {}, False

    blocks = []
    current_block = [team_occurrences[0]]
    for i in range(1, len(team_occurrences)):
        row_idx, val = team_occurrences[i]
        prev_row = current_block[-1][0]
        gap = row_idx - prev_row
        if gap <= 1 or (gap <= 3 and len(current_block) < len(TEAMS)):
            current_block.append((row_idx, val))
        else:
            blocks.append(current_block)
            current_block = [(row_idx, val)]
    blocks.append(current_block)

    def _expand_block(block):
        if not block:
            return {}
        start_row = block[0][0]
        end_row = block[-1][0]
        result = {}
        row_idx = start_row
        team_pos = 0
        expected_teams = list(TEAMS)
        while row_idx <= end_row and team_pos < len(expected_teams):
            val = ws.cell(row=row_idx, column=3).value
            if val in TEAMS:
                result[val] = row_idx
                team_pos += 1
            elif val is None or (isinstance(val, str) and val.strip() == ""):
                result[f"<missing>@{row_idx}"] = row_idx
            row_idx += 1
        return result

    new_rows = {}
    followup_rows = {}
    fu_present = False

    if len(blocks) >= 2:
        new_rows = _expand_block(blocks[0])
        followup_rows = _expand_block(blocks[1])
        fu_present = True
    elif len(blocks) == 1:
        new_rows = _expand_block(blocks[0])

    return new_rows, followup_rows, fu_present


def _detect_secondary_row(ws, new_rows):
    """Find the secondary metrics data row by scanning for 'Unrecognised'
    in column B (the secondary block header). Data is always 2 rows below."""
    max_case_row = max(list(new_rows.values()) + [0])
    for row_idx in range(max_case_row + 1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=2).value
        if isinstance(val, str) and "Unrecogn" in val:
            return row_idx + 2
    return max_case_row + 4


# ---------------------------------------------------------------------------
# Main parse entry point
# ---------------------------------------------------------------------------

def parse_excel(path_or_bytes, file_hash: str = "") -> ParsedReport:
    """Parse a weekly or monthly Wellness report workbook.

    Raises SheetStructureError if the workbook is not a Wellness report at all.
    """
    if isinstance(path_or_bytes, (bytes, bytearray)):
        wb_values = openpyxl.load_workbook(io.BytesIO(bytes(path_or_bytes)), data_only=True)
        wb_formulas = openpyxl.load_workbook(io.BytesIO(bytes(path_or_bytes)), data_only=False)
    else:
        wb_values = openpyxl.load_workbook(path_or_bytes, data_only=True)
        wb_formulas = openpyxl.load_workbook(path_or_bytes, data_only=False)

    if not file_hash and isinstance(path_or_bytes, str):
        with open(path_or_bytes, "rb") as fh:
            file_hash = hashlib.sha256(fh.read()).hexdigest()
    elif not file_hash and isinstance(path_or_bytes, (bytes, bytearray)):
        file_hash = hashlib.sha256(bytes(path_or_bytes)).hexdigest()

    ws = wb_values.active
    wsf = wb_formulas.active

    title = None
    title_cell = None
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if isinstance(val, str) and val.strip():
            title = val
            title_cell = f"{openpyxl.utils.get_column_letter(col_idx)}1"
            break
    if title is None:
        raise SheetStructureError(ERR_SHEET_STRUCTURE, "No title string found in row 1.")

    parsed_title = parse_title(title)
    if parsed_title is None:
        raise SheetStructureError(
            ERR_SHEET_STRUCTURE,
            'expected title "Weekly/Monthly Wellness Report From <start> to <end> <year>" '
            f"in row 1 (found {title!r} in cell {title_cell})",
        )
    report_type, period_start, period_end = parsed_title

    # --- detect actual row layout ---
    new_rows, followup_rows, fu_present = _detect_team_rows(ws)

    # --- structural guards (spec section 15) ---
    structure_issues = []
    if "WLN Ctr" not in new_rows:
        structure_issues.append("WLN Ctr sub-team row not found in the sheet")

    structure_ok = not structure_issues

    # --- case rows ---
    rows: list[SubTeamRow] = []
    for team, row_num in new_rows.items():
        if team.startswith("<missing>"):
            rows.append(
                SubTeamRow(
                    case_type="new",
                    sub_team="<missing>",
                    sheet_row=row_num,
                    columns={name: 0 for name in COLUMN_NAMES},
                    status="rejected",
                    issues=[CellIssue("", f"C{row_num}", ERR_MISSING_MANDATORY,
                                      f"Row {row_num}: sub-team row missing (expected one of {TEAMS}).")],
                )
            )
            continue
        row_obj = _parse_case_row(ws, wsf, "new", team, row_num)
        rows.append(row_obj)
    for team, row_num in followup_rows.items():
        if team.startswith("<missing>"):
            rows.append(
                SubTeamRow(
                    case_type="followup",
                    sub_team="<missing>",
                    sheet_row=row_num,
                    columns={name: 0 for name in COLUMN_NAMES},
                    status="rejected",
                    issues=[CellIssue("", f"C{row_num}", ERR_MISSING_MANDATORY,
                                      f"Row {row_num}: sub-team row missing (expected one of {TEAMS}).")],
                )
            )
            continue
        row_obj = _parse_case_row(ws, wsf, "followup", team, row_num)
        rows.append(row_obj)

    # --- secondary metrics block ---
    sec_row = _detect_secondary_row(ws, new_rows)
    secondary = _parse_secondary(ws, wsf, sec_row)

    warnings = []
    if structure_issues:
        warnings.append(f"Layout warnings: {'; '.join(structure_issues)}")
    if not fu_present:
        warnings.append("This file only contains new-case data - the period will be marked incomplete.")
    if secondary.stray_cells:
        stray_vals = {k: v for k, v in secondary.stray_cells.items() if v}
        if stray_vals:
            warnings.append(f"Legacy stray cells detected in the secondary block (ignored): {stray_vals}")

    return ParsedReport(
        report_type=report_type,
        period_start=period_start,
        period_end=period_end,
        title=title,
        rows=rows,
        secondary=secondary,
        structure_ok=structure_ok,
        structure_issues=structure_issues,
        warnings=warnings,
        title_range_mismatch=_range_mismatch(report_type, period_start, period_end),
        file_sha256=file_hash,
    )


def _range_mismatch(report_type, start, end) -> bool:
    days = (end - start).days + 1
    if report_type == "weekly":
        return days not in (6, 7, 8)
    return days <= 7


def _parse_case_row(ws, wsf, case_type: str, sub_team: str, row_num: int) -> SubTeamRow:
    columns: dict = {}
    issues: list[CellIssue] = []
    for i, (field, _group) in enumerate(COLUMN_SCHEMA):
        col_idx = DATA_START_COL + i
        cell_ref = f"{openpyxl.utils.get_column_letter(col_idx)}{row_num}"
        value = _cell_value(ws, wsf, row_num, col_idx)
        parsed, issue = _coerce_int(value, cell_ref, field)
        if issue is not None:
            issues.append(issue)
            parsed = 0
        columns[field] = parsed

    checks = run_checks(columns)

    if issues:
        return SubTeamRow(
            case_type=case_type,
            sub_team=sub_team,
            sheet_row=row_num,
            columns=columns,
            checks=checks,
            status="rejected",
            issues=issues,
        )
    if any(not c.passed for c in checks):
        return SubTeamRow(
            case_type=case_type,
            sub_team=sub_team,
            sheet_row=row_num,
            columns=columns,
            checks=checks,
            status="warning",
        )
    return SubTeamRow(
        case_type=case_type,
        sub_team=sub_team,
        sheet_row=row_num,
        columns=columns,
        checks=checks,
        status="ready",
    )


def _cell_value(ws, wsf, row: int, col: int):
    """Return the cached value if present; otherwise evaluate a simple =SUM()
    formula from the formula workbook over the values workbook."""
    value = ws.cell(row=row, column=col).value
    if value is not None:
        return value
    formula = wsf.cell(row=row, column=col).value
    if isinstance(formula, str) and formula.startswith("="):
        refs = _sum_formula_range(formula)
        if refs:
            total = 0
            ok = True
            for ref in refs:
                import re as _re

                mm = _re.match(r"([A-Z]+)(\d+)", ref)
                c = openpyxl.utils.column_index_from_string(mm.group(1))
                r = int(mm.group(2))
                val = ws.cell(row=r, column=c).value
                if val is None or not isinstance(val, (int, float)):
                    ok = False
                    break
                total += val
            if ok:
                return total
    return value


def run_checks(columns: dict) -> list[CheckResult]:
    total = columns.get("total_cases", 0)
    results = []
    for name, fields in CHECK_GROUPS.items():
        actual = sum(columns.get(f, 0) for f in fields)
        results.append(CheckResult(name=name, passed=actual == total, expected=total, actual=actual))
    return results


def merge_verticals(rows) -> dict:
    """Normalize source rows into the four independent report verticals."""
    merged: dict = {}
    for case_type in ("new", "followup"):
        accepted = [r for r in rows if r.case_type == case_type and r.status != "rejected"]
        by_team = {r.sub_team: r for r in accepted}
        for vertical, members in {
            "WC": ["WLN Ctr"],
            "TA": ["Team A"],
            "YD": ["Your Dost"],
            "MW": ["Myndwell"],
        }.items():
            row: dict = {}
            for field in COLUMN_NAMES:
                row[field] = sum(by_team[t].columns[field] for t in members if t in by_team)
            merged[(case_type, vertical)] = row
    return merged


def _parse_secondary(ws, wsf, sec_row=20) -> SecondaryMetrics:
    sec = SecondaryMetrics()
    team_cells = {"WLN Ctr": 0, "Team A": 1, "Your Dost": 2, "Myndwell": 3}

    for group_name, (header_col, start_col, total_col) in SECONDARY_GROUPS.items():
        teams_raw = [
            _cell_value(ws, wsf, sec_row, start_col + i) or 0 for i in range(4)
        ]
        total = _cell_value(ws, wsf, sec_row, total_col)
        if total is None:
            total = sum(teams_raw)
        sec_dest = getattr(sec, group_name)
        sec_dest["WC"] = teams_raw[0]
        sec_dest["TA"] = teams_raw[1]
        sec_dest["YD"] = teams_raw[2]
        sec_dest["MW"] = teams_raw[3]
        sec_dest["Total"] = int(total)

    for key, col in ENQUIRY_COLS.items():
        sec.enquiry_modes[key] = int(_cell_value(ws, wsf, sec_row, col) or 0)

    for i, col in enumerate(STRAY_COLS):
        label = ws.cell(row=sec_row - 1, column=col).value
        val = _cell_value(ws, wsf, sec_row, col)
        if val is not None and val != 0:
            sec.stray_cells[f"{col_to_letter(col)}{sec_row} ({label or '?'})"] = val
    return sec


def col_to_letter(col: int) -> str:
    return openpyxl.utils.get_column_letter(col)


class SheetStructureError(Exception):
    def __init__(self, code: str, message: str):
        self.code = code
        self.message = message
        super().__init__(message)
