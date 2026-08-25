import io

import openpyxl
from django.db import migrations


def rebuild_secondary_verticals(apps, schema_editor):
    Period = apps.get_model("wellness", "Period")
    SecondaryMetrics = apps.get_model("wellness", "SecondaryMetrics")
    EnquiryModes = apps.get_model("wellness", "EnquiryModes")
    ImportEvent = apps.get_model("wellness", "ImportEvent")

    groups = {
        "total_sessions": (4, 8),
        "early_prevention_warning": (9, 13),
        "no_show_turn_up": (14, 18),
        "active_cases": (19, 23),
        "clients_over_4_sessions": (24, 28),
    }

    def integer(value):
        return int(value or 0)

    for period in Period.objects.all().iterator():
        event = ImportEvent.objects.filter(period_id=period.id).order_by("-id").first()
        if event is None or not event.file:
            continue
        try:
            with event.file.open("rb") as source:
                ws = openpyxl.load_workbook(io.BytesIO(source.read()), data_only=True).active
        except Exception:
            continue

        values = {name: [integer(ws.cell(20, start + i).value) for i in range(4)] for name, (start, _total) in groups.items()}
        total_values = {name: integer(ws.cell(20, total).value) for name, (_start, total) in groups.items()}
        for index, vertical in enumerate(("WC", "TA", "YD", "MW")):
            SecondaryMetrics.objects.update_or_create(
                period_id=period.id,
                vertical=vertical,
                defaults={name: values[name][index] for name in groups},
            )
        SecondaryMetrics.objects.update_or_create(
            period_id=period.id,
            vertical="Total",
            defaults=total_values,
        )
        EnquiryModes.objects.update_or_create(
            period_id=period.id,
            defaults={
                "mail": integer(ws.cell(20, 29).value),
                "calls_recd": integer(ws.cell(20, 30).value),
                "calls_out": integer(ws.cell(20, 31).value),
            },
        )


class Migration(migrations.Migration):
    dependencies = [("wellness", "0002_separate_team_a")]
    operations = [migrations.RunPython(rebuild_secondary_verticals, migrations.RunPython.noop)]
