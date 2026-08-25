"""Helpers to build synthetic Wellness-report workbooks for tests."""

import openpyxl

from wellness.services.parsing import COLUMN_SCHEMA, TEAMS, DATA_START_COL


def fields(**kwargs):
    """A full 29-field dict, defaulting to 0."""
    f = {name: 0 for name, _ in COLUMN_SCHEMA}
    f.update(kwargs)
    return f


def _write_row(ws, row, values: dict):
    for i, (name, _g) in enumerate(COLUMN_SCHEMA):
        v = values.get(name)
        if v is not None:
            ws.cell(row=row, column=DATA_START_COL + i, value=v)
    ws.cell(row=row, column=2, value="No. of New Cases" if row < 10 else "No. of Follow-up cases")


def build_workbook(
    report_type="weekly",
    start="29th July",
    end="04th August 2026",
    new_rows=None,
    fu_rows=None,
    secondary=None,
    include_followup=True,
    subteam_labels=None,
):
    """Build a synthetic workbook matching the fixed layout.

    new_rows / fu_rows: dict {team: {field: int}}.
    subteam_labels: optional override for the C-column labels (to simulate
        malformed rows).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B1"] = f"{report_type.capitalize()} Wellness Report From {start} to {end}"
    ws["C6"] = "WLN Ctr"
    ws["B10"] = "Total no. of cases NEW"
    ws["B16"] = "Grand Total"
    ws["B18"] = "Unrecognised"

    new_rows = new_rows or {t: fields() for t in TEAMS}
    for i, team in enumerate(TEAMS):
        row = 6 + i
        label = team
        if subteam_labels:
            label = subteam_labels.get(team, team)
        ws.cell(row=row, column=3, value=label)
        _write_row(ws, row, new_rows.get(team, fields()))

    if include_followup:
        fu_rows = fu_rows or {t: fields() for t in TEAMS}
        ws["C11"] = "WLN Ctr"
        ws["B15"] = "Total no. of cases FOLLOW-UP"
        for i, team in enumerate(TEAMS):
            row = 11 + i
            label = team
            if subteam_labels:
                label = subteam_labels.get(team, team)
            ws.cell(row=row, column=3, value=label)
            _write_row(ws, row, fu_rows.get(team, fields()))

    if secondary:
        # rows 19-20: teams in order WLN Ctr, Team A, Your Dost, Myndwell
        for group, (header_col, start_col) in {
            "total_sessions": (18, 4),
            "early_prevention_warning": (18, 9),
            "no_show_turn_up": (18, 14),
            "active_cases": (18, 19),
            "clients_over_4_sessions": (18, 24),
        }.items():
            vals = secondary.get(group, {})
            for i, team in enumerate(TEAMS):
                ws.cell(row=20, column=start_col + i, value=int(vals.get(team, 0)))
            ws.cell(row=20, column=start_col + 4, value=int(vals.get("Total", 0)))
        modes = secondary.get("enquiry_modes", {})
        ws.cell(row=20, column=29, value=int(modes.get("mail", 0)))
        ws.cell(row=20, column=30, value=int(modes.get("calls_recd", 0)))
        ws.cell(row=20, column=31, value=int(modes.get("calls_out", 0)))
    return wb
