import openpyxl
import json

wb = openpyxl.load_workbook(r'C:\Users\Wellness\Downloads\29th July to 04th August and 05th to 11th August 2026.xlsx')

# Check all sheets for charts
for name in wb.sheetnames:
    ws = wb[name]
    if ws._charts:
        print(f'\n=== Charts in {name}: {len(ws._charts)} ===')
        for c in ws._charts:
            ctype = type(c).__name__
            print(f'  {ctype}: title="{c.title}", anchor={c.anchor}')
            if hasattr(c, 'series'):
                for s in c.series:
                    sname = str(s.title) if s.title else 'N/A'
                    print(f'    Series: {sname}')
            if hasattr(c, 'categories'):
                print(f'    Categories: {c.categories}')
    else:
        print(f'No charts in {name}')

# Check conditional formatting, fills, font colors on Report sheet
ws = wb['Report']
print('\n=== REPORT FORMATTING ===')
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        if cell.value is not None:
            fill_color = None
            font_color = None
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != '00000000':
                fill_color = cell.fill.fgColor.rgb
            if cell.font and cell.font.color and cell.font.color.rgb:
                font_color = cell.font.color.rgb
            if fill_color or font_color:
                print(f'  {cell.coordinate}: val={cell.value}, fill={fill_color}, font_color={font_color}, bold={cell.font.bold if cell.font else None}')

# Get merged cells
print('\n=== MERGED CELLS (Report) ===')
for mc in ws.merged_cells.ranges:
    print(f'  {mc}')

# Row heights and col widths
print('\n=== COLUMN WIDTHS (Report) ===')
for col_letter, dim in ws.column_dimensions.items():
    if dim.width and dim.width != 8:
        print(f'  {col_letter}: width={dim.width}')

print('\n=== ROW HEIGHTS (Report) ===')
for row_num, dim in ws.row_dimensions.items():
    if dim.height and dim.height != 15:
        print(f'  Row {row_num}: height={dim.height}')
